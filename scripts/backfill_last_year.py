#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import csv
import io
import json
import os
import re
from datetime import date
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

CITY_PAGE = "https://www.city.yokohama.lg.jp/kosodate-kyoiku/hoiku-yoji/shisetsu/riyou/info/nyusho-jokyo.html"

WARD_FILTER = (os.getenv("WARD_FILTER", "") or "").strip() or None
MONTHS_BACK = int(os.getenv("MONTHS_BACK", "24"))
FORCE = (os.getenv("FORCE_BACKFILL", "0") == "1")

# master を backfill 時点で注入する（1推奨）
APPLY_MASTER = (os.getenv("APPLY_MASTER", "1") == "1")

ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
DATA_DIR.mkdir(parents=True, exist_ok=True)
MASTER_CSV = DATA_DIR / "master_facilities.csv"
MONTHS_JSON = DATA_DIR / "months.json"


# ---------- small utils ----------
def norm(s: Any) -> str:
    if s is None:
        return ""
    x = str(s).replace("　", " ")
    x = re.sub(r"\s+", "", x)
    return x.strip()


def safe(x: Any) -> str:
    return "" if x is None else str(x)


def to_int(x: Any) -> Optional[int]:
    if x is None:
        return None
    s = str(x).strip()
    if s == "" or s.lower() == "nan":
        return None
    if s in ("-", "－", "‐", "—", "―"):
        return 0
    try:
        return int(float(s))
    except Exception:
        return None


def sum_opt(*vals: Optional[int]) -> Optional[int]:
    xs = [v for v in vals if v is not None]
    return sum(xs) if xs else None


def ratio_opt(wait: Optional[int], cap: Optional[int]) -> Optional[float]:
    if wait is None or cap in (None, 0):
        return None
    return wait / cap


def month_floor(d: date) -> date:
    return date(d.year, d.month, 1)


def add_months(d: date, delta: int) -> date:
    y, m = d.year, d.month
    m2 = m + delta
    y += (m2 - 1) // 12
    m2 = (m2 - 1) % 12 + 1
    return date(y, m2, 1)


def iso(d: date) -> str:
    return d.isoformat()


def sanitize_header(header: List[str]) -> List[str]:
    out = []
    seen: Dict[str, int] = {}
    for i, h in enumerate(header):
        h2 = (h or "").strip()
        if h2 == "":
            h2 = f"col{i}"
        if h2 in seen:
            seen[h2] += 1
            h2 = f"{h2}_{seen[h2]}"
        else:
            seen[h2] = 0
        out.append(h2)
    return out


def extract_month_from_text(text: str) -> Optional[str]:
    """
    例: '【令和８年２月１日時点】' → 2026-02-01
    """
    if not text:
        return None
    t = str(text)
    z2h = str.maketrans("０１２３４５６７８９", "0123456789")
    t = t.translate(z2h)

    m = re.search(r"令和\s*([0-9]+)\s*年\s*([0-9]+)\s*月\s*1\s*日", t)
    if m:
        ry = int(m.group(1))
        mm = int(m.group(2))
        y = 2018 + ry  # Reiwa 1 = 2019
        return date(y, mm, 1).isoformat()

    m = re.search(r"([0-9]{4})\s*年\s*([0-9]{1,2})\s*月\s*1\s*日", t)
    if m:
        y = int(m.group(1))
        mm = int(m.group(2))
        return date(y, mm, 1).isoformat()

    return None


def detect_month_from_rows(rows: List[Dict[str, str]]) -> Optional[str]:
    if not rows:
        return None
    for k in ("更新日", "更新年月日", "更新日時", "更新年月"):
        v = str(rows[0].get(k, "")).strip()
        if v:
            v = v[:10].replace("/", "-")
            try:
                y, m, _ = v.split("-")
                return date(int(y), int(m), 1).isoformat()
            except Exception:
                return None
    return None


def extract_mm_only(text: str) -> Optional[int]:
    """
    '4月', '04月', '４月', '（4月1日）' などから月だけ抜く
    """
    if not text:
        return None
    t = str(text)
    z2h = str.maketrans("０１２３４５６７８９", "0123456789")
    t = t.translate(z2h)
    m = re.search(r"\b([0-9]{1,2})\s*月\b", t)
    if not m:
        m = re.search(r"^([0-9]{1,2})\s*月", t)
    if m:
        mm = int(m.group(1))
        if 1 <= mm <= 12:
            return mm
    return None


# ---------- master apply ----------
def load_master() -> Dict[str, Dict[str, str]]:
    if not MASTER_CSV.exists():
        return {}
    out: Dict[str, Dict[str, str]] = {}
    with MASTER_CSV.open("r", encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            fid = safe(row.get("facility_id")).strip()
            if fid:
                out[fid] = {k: safe(v) for k, v in row.items()}
    return out


def as_int_str(x: Any) -> Optional[str]:
    s = safe(x).strip()
    if s == "" or s.lower() == "null" or s == "-":
        return None
    try:
        return str(int(float(s)))
    except Exception:
        return None


def apply_master_to_facility(f: Dict[str, Any], m: Dict[str, str]) -> int:
    """
    master に値がある項目だけ注入する（空で上書きしない）
    """
    updated = 0
    mapping = {
        "address": "address",
        "lat": "lat",
        "lng": "lng",
        "map_url": "map_url",
        "facility_type": "facility_type",
        "phone": "phone",
        "website": "website",
        "notes": "notes",
        "nearest_station": "nearest_station",
        "name_kana": "name_kana",
        "station_kana": "station_kana",
    }
    for jkey, mkey in mapping.items():
        mv = safe(m.get(mkey)).strip()
        if mv == "":
            continue
        cur = safe(f.get(jkey)).strip()
        if cur != mv:
            f[jkey] = mv
            updated += 1

    wm = as_int_str(m.get("walk_minutes"))
    if wm is not None:
        cur = safe(f.get("walk_minutes")).strip()
        if cur != wm:
            f["walk_minutes"] = wm
            updated += 1

    return updated


# ---------- scraping ----------
def parse_reiwa_fy(text: str) -> Optional[int]:
    """
    '令和6年度' -> 2024 を返す（年度開始年）
    """
    if not text:
        return None
    t = str(text)
    z2h = str.maketrans("０１２３４５６７８９", "0123456789")
    t = t.translate(z2h)
    m = re.search(r"令和\s*([0-9]{1,2})\s*年度", t)
    if not m:
        return None
    ry = int(m.group(1))
    return 2018 + ry  # 令和1年度=2019年度（開始年2019）


def scrape_excel_urls() -> Dict[str, List[Dict[str, Any]]]:
    """
    横浜市ページから Excel リンクを拾って分類する（年度情報も付与）
    返り値例:
      {"accept":[{"url":..., "fy":2025}, ...], "wait":[...], "enrolled":[...]}
    """
    html = requests.get(CITY_PAGE, timeout=30).text
    soup = BeautifulSoup(html, "html.parser")

    # 文書順に h2/h3/h4 と a を走査し、直近の「令和◯年度」を current_fy として持つ
    current_fy: Optional[int] = None
    found: List[Tuple[str, str, Optional[int]]] = []

    for el in soup.select("h1,h2,h3,h4,a[href]"):
        if el.name in ("h1", "h2", "h3", "h4"):
            fy = parse_reiwa_fy(el.get_text(" ", strip=True))
            if fy:
                current_fy = fy
            continue

        if el.name == "a":
            text = el.get_text(" ", strip=True)
            href = (el.get("href") or "").strip()
            if not href:
                continue

            href_abs = href if href.startswith("http") else requests.compat.urljoin(CITY_PAGE, href)

            # 拾い方を拡張：
            # - URLに拡張子がなくても、テキストに「エクセル」があり、かつ目的語（受入/待ち/児童）があれば拾う
            href_l = href_abs.lower()
            is_excel_like = any(x in href_l for x in (".xlsx", ".xlsm", ".xls")) or ("エクセル" in text) or ("excel" in text.lower())

            if not is_excel_like:
                continue

            if any(k in text for k in ("入所児童", "受入可能", "入所待ち", "待ち人数", "入所状況")):
                found.append((href_abs, text, current_fy))

    # HTML直書きURLも拾う（年度は不明なのでNone）
    for u in re.findall(r"https?://[^\s\"']+\.(?:xlsx|xlsm|xls)(?:\?[^\s\"']*)?", html, flags=re.I):
        found.append((u, "", None))

    # uniq
    seen = set()
    uniq: List[Tuple[str, str, Optional[int]]] = []
    for u, t, fy in found:
        key = (u, fy)
        if key not in seen:
            seen.add(key)
            uniq.append((u, t, fy))

    urls: Dict[str, List[Dict[str, Any]]] = {"accept": [], "wait": [], "enrolled": []}

    def add(kind: str, u: str, t: str, fy: Optional[int]):
        urls[kind].append({"url": u, "text": t, "fy": fy})

    for u, t, fy in uniq:
        if "入所児童" in t:
            add("enrolled", u, t, fy)
        elif "受入可能" in t:
            add("accept", u, t, fy)
        elif ("入所待ち" in t) or ("待ち人数" in t):
            add("wait", u, t, fy)

    # フォールバック（テキスト分類できない場合はURL文字列で当てる）
    def push_if(kind: str, pred):
        if urls[kind]:
            return
        for u, t, fy in uniq:
            ul = u.lower()
            if pred(ul):
                add(kind, u, t, fy)

    push_if("accept",   lambda ul: ("0932_" in ul) or ("ukire" in ul) or ("受入" in ul))
    push_if("wait",     lambda ul: ("0933_" in ul) or ("mati" in ul) or ("待ち" in ul) or ("0929_" in ul))
    push_if("enrolled", lambda ul: ("0934_" in ul) or ("jido" in ul) or ("児童" in ul) or ("0923_" in ul))

    # sort: fy昇順→URL昇順（再現性）
    for k in urls.keys():
        urls[k].sort(key=lambda d: (d.get("fy") or 0, d.get("url") or ""))

    if not urls["accept"] or not urls["wait"]:
        sample = [d["url"] for d in urls["accept"][:5]] + [d["url"] for d in urls["wait"][:5]]
        raise RuntimeError(f"Excelリンクが拾えません（accept={len(urls['accept'])}, wait={len(urls['wait'])}, sample={sample}）")

    print("XLS links found:", {k: len(v) for k, v in urls.items()})
    return urls


# ---------- Excel parsing ----------
def sheet_to_rows(ws) -> List[List[Any]]:
    rows: List[List[Any]] = []
    max_r = min(ws.max_row or 0, 6000)
    max_c = min(ws.max_column or 0, 120)
    for r in range(1, max_r + 1):
        row = []
        for c in range(1, max_c + 1):
            row.append(ws.cell(r, c).value)
        rows.append(row)
    return rows


def find_header_index(rows: List[List[Any]]) -> Optional[int]:
    keywords = ("施設", "区", "合計", "0歳", "０歳", "1歳", "１歳", "受入", "待ち", "児童")
    best_i, best_score = None, -1
    for i, row in enumerate(rows[:120]):
        cells = ["" if v is None else str(v) for v in row]
        nonempty = sum(1 for c in cells if c.strip() != "")
        has_kw = any(any(k in c for k in keywords) for c in cells)
        score = nonempty + (10 if has_kw else 0)
        if nonempty >= 5 and score > best_score:
            best_i, best_score = i, score
    return best_i


def month_from_fy(mm: int, fy_start_year: int) -> str:
    """
    FY(4月開始)に基づき、mm(1-12)の年月を確定してISO返す
    """
    y = fy_start_year if mm >= 4 else (fy_start_year + 1)
    return date(y, mm, 1).isoformat()


def parse_sheet(ws, fy_start_year: Optional[int]) -> Tuple[Optional[str], List[Dict[str, str]]]:
    rows = sheet_to_rows(ws)

    # 1) まずは明示（令和/西暦）を最優先
    month = extract_month_from_text(ws.title)
    if month is None:
        for r in rows[:20]:
            for v in r[:10]:
                month = extract_month_from_text("" if v is None else str(v))
                if month:
                    break
            if month:
                break

    hidx = find_header_index(rows)
    if hidx is None:
        return month, []

    header = sanitize_header([("" if v is None else str(v)) for v in rows[hidx]])
    out: List[Dict[str, str]] = []

    empty_streak = 0
    for r in rows[hidx + 1 :]:
        vals = [("" if v is None else str(v)) for v in r]
        if all(v.strip() == "" for v in vals):
            empty_streak += 1
            if empty_streak >= 10:
                break
            continue
        empty_streak = 0
        d = {header[i]: vals[i] if i < len(vals) else "" for i in range(len(header))}
        out.append(d)

    # 2) 行の更新日から取れるなら採用
    m2 = detect_month_from_rows(out)
    if m2:
        month = m2

    # 3) それでも年が取れない場合：シート名/先頭セル等から “月だけ” を拾い、FYで確定
    if month is None and fy_start_year:
        mm = extract_mm_only(ws.title)
        if mm is None:
            # 先頭付近の文字列にも “4月” があるケース
            for r in rows[:10]:
                for v in r[:6]:
                    mm = extract_mm_only("" if v is None else str(v))
                    if mm:
                        break
                if mm:
                    break
        if mm:
            month = month_from_fy(mm, fy_start_year)

    return month, out


def read_xlsx(url: str, fy_start_year: Optional[int]) -> Dict[str, List[Dict[str, str]]]:
    """
    xlsx 1ファイル → {month: rows} を返す（同月が複数シートなら後勝ち）
    """
    print("download:", url, "fy_start_year:", fy_start_year)
    r = requests.get(url, timeout=120)
    r.raise_for_status()
    wb = load_workbook(io.BytesIO(r.content), data_only=True)

    mp: Dict[str, List[Dict[str, str]]] = {}
    for ws in wb.worksheets:
        month, rows = parse_sheet(ws, fy_start_year=fy_start_year)
        if month and rows:
            mp[month] = rows

    if mp:
        rng = (sorted(mp.keys())[0], sorted(mp.keys())[-1])
        print("  parsed months:", len(mp), "range:", rng)
    else:
        print("  parsed months: 0")
    return mp


# ---------- column guessing / metrics ----------
def guess_facility_id_key(rows: List[Dict[str, str]]) -> str:
    header = list(rows[0].keys())

    candidates = [
        "施設番号", "施設・事業所番号", "施設事業所番号", "事業所番号",
        "施設ID", "施設ＩＤ", "施設・事業所ID", "施設・事業所ＩＤ",
        "施設No", "施設Ｎｏ", "事業所No", "事業所Ｎｏ",
    ]
    for k in candidates:
        if k in rows[0]:
            return k

    patterns = ("番号", "ID", "ＩＤ", "No", "Ｎｏ", "NO", "ＮＯ")
    for k in header:
        if any(p in k for p in patterns) and ("施設" in k or "事業所" in k):
            return k

    N = min(200, len(rows))
    digit_re = re.compile(r"^\d{4,}$")
    best_key, best_score = None, -1
    for k in header:
        score = 0
        for i in range(N):
            v = str(rows[i].get(k, "")).strip()
            if digit_re.match(v):
                score += 1
        if score > best_score:
            best_key, best_score = k, score

    if best_key and best_score >= max(10, int(N * 0.30)):
        return best_key

    raise RuntimeError("施設番号列が見つかりません")


def index_by_key(rows: List[Dict[str, str]], key: str) -> Dict[str, Dict[str, str]]:
    out: Dict[str, Dict[str, str]] = {}
    for r in rows:
        v = str(r.get(key, "")).strip()
        if v:
            out[v] = r
    return out


def pick_ward_key(row: Dict[str, str]) -> Optional[str]:
    for k in ("施設所在区", "所在区", "区名"):
        if k in row:
            return k
    for k in row.keys():
        if "区" in k:
            return k
    return None


def pick_name_key(row: Dict[str, str]) -> Optional[str]:
    for k in ("施設名", "施設・事業名", "施設・事業所名", "事業名"):
        if k in row:
            return k
    for k in row.keys():
        if "施設" in k and "区" not in k:
            return k
    return None


def get_total(row: Dict[str, str]) -> Optional[int]:
    if not row:
        return None
    if "合計" in row and str(row.get("合計", "")).strip() != "":
        return to_int(row.get("合計"))
    for k in row.keys():
        if "合計" in k and str(row.get(k, "")).strip() != "":
            return to_int(row.get(k))
    return None


def get_age_value(row: Dict[str, str], age: int) -> Optional[int]:
    if not row:
        return None
    z = "０１２３４５"
    pats = [f"{age}歳児", f"{age}歳", z[age] + "歳児", z[age] + "歳"]
    for p in pats:
        if p in row and str(row.get(p, "")).strip() != "":
            return to_int(row.get(p))
    for k in row.keys():
        if any(p in k for p in pats) and str(row.get(k, "")).strip() != "":
            return to_int(row.get(k))
    return None


def build_age_groups(ar: Dict[str, str], wr: Dict[str, str], er: Dict[str, str]) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    ages_0_5: Dict[str, Dict[str, Any]] = {}
    for i in range(6):
        a = get_age_value(ar, i)
        w = get_age_value(wr, i) if wr else None
        e = get_age_value(er, i) if er else None
        cap_est = (e + a) if (e is not None and a is not None) else None
        ages_0_5[str(i)] = {
            "accept": a,
            "wait": w,
            "enrolled": e,
            "capacity_est": cap_est,
            "wait_per_capacity_est": ratio_opt(w, cap_est),
        }

    g0, g1, g2 = ages_0_5["0"], ages_0_5["1"], ages_0_5["2"]
    g3, g4, g5 = ages_0_5["3"], ages_0_5["4"], ages_0_5["5"]

    w_35 = sum_opt(g3.get("wait"), g4.get("wait"), g5.get("wait"))
    cap_35 = sum_opt(g3.get("capacity_est"), g4.get("capacity_est"), g5.get("capacity_est"))

    age_groups = {
        "0": g0,
        "1": g1,
        "2": g2,
        "3-5": {
            "accept": sum_opt(g3.get("accept"), g4.get("accept"), g5.get("accept")),
            "wait": w_35,
            "enrolled": sum_opt(g3.get("enrolled"), g4.get("enrolled"), g5.get("enrolled")),
            "capacity_est": cap_35,
            "wait_per_capacity_est": ratio_opt(w_35, cap_35),
        },
    }
    return age_groups, ages_0_5


# ---------- main backfill ----------
def main() -> None:
    print("BACKFILL start. ward=", WARD_FILTER, "months_back=", MONTHS_BACK, "force=", FORCE, "apply_master=", APPLY_MASTER)

    urls = scrape_excel_urls()
    master = load_master() if APPLY_MASTER else {}
    target = norm(WARD_FILTER) if WARD_FILTER else None

    acc_by_month: Dict[str, List[Dict[str, str]]] = {}
    wai_by_month: Dict[str, List[Dict[str, str]]] = {}
    enr_by_month: Dict[str, List[Dict[str, str]]] = {}

    # 重要：年度(fy)つきで読む
    for d in urls["accept"]:
        try:
            acc_by_month.update(read_xlsx(d["url"], fy_start_year=d.get("fy")))
        except Exception as e:
            print("WARN accept xlsx failed:", d.get("url"), e)

    for d in urls["wait"]:
        try:
            wai_by_month.update(read_xlsx(d["url"], fy_start_year=d.get("fy")))
        except Exception as e:
            print("WARN wait xlsx failed:", d.get("url"), e)

    for d in urls["enrolled"]:
        try:
            enr_by_month.update(read_xlsx(d["url"], fy_start_year=d.get("fy")))
        except Exception as e:
            print("WARN enrolled xlsx failed:", d.get("url"), e)

    if not acc_by_month:
        raise RuntimeError("受入可能数の月次が1つも取れませんでした")

    end = month_floor(date.today())
    start = add_months(end, -(MONTHS_BACK - 1))
    want: List[str] = []
    cur = start
    while cur <= end:
        want.append(iso(cur))
        cur = add_months(cur, 1)

    available = [m for m in want if m in acc_by_month]
    print("want months:", len(want), "available:", len(available), "missing:", [m for m in want if m not in acc_by_month][:30], "...")


    existing_months: List[str] = []
    if MONTHS_JSON.exists():
        try:
            existing_months = json.loads(MONTHS_JSON.read_text(encoding="utf-8")).get("months", [])
        except Exception:
            existing_months = []

    changed_any = 0

    for m in available:
        out_path = DATA_DIR / f"{m}.json"
        if out_path.exists() and not FORCE:
            print("skip exists:", out_path.name)
            continue

        accept_rows = acc_by_month.get(m, [])
        wait_rows = wai_by_month.get(m, [])
        enrolled_rows = enr_by_month.get(m, [])

        fid_a = guess_facility_id_key(accept_rows)
        A = index_by_key(accept_rows, fid_a)

        W = {}
        if wait_rows:
            try:
                fid_w = guess_facility_id_key(wait_rows)
                W = index_by_key(wait_rows, fid_w)
            except Exception:
                W = {}

        E = {}
        if enrolled_rows:
            try:
                fid_e = guess_facility_id_key(enrolled_rows)
                E = index_by_key(enrolled_rows, fid_e)
            except Exception:
                E = {}

        ward_key = pick_ward_key(accept_rows[0]) if accept_rows else None
        name_key = pick_name_key(accept_rows[0]) if accept_rows else None

        facilities: List[Dict[str, Any]] = []
        injected_cells = 0

        for fid, ar in A.items():
            ward = norm(ar.get(ward_key)) if ward_key else ""
            ward = ward.replace("横浜市", "")
            if target and target not in ward:
                continue

            wr = W.get(fid, {})
            er = E.get(fid, {})

            name = str(ar.get(name_key, "")).strip() if name_key else ""

            tot_accept = get_total(ar)
            tot_wait = get_total(wr) if wr else None
            tot_enrolled = get_total(er) if er else None
            cap_est = (tot_enrolled + tot_accept) if (tot_enrolled is not None and tot_accept is not None) else None

            age_groups, ages_0_5 = build_age_groups(ar, wr, er)

            fobj: Dict[str, Any] = {
                "id": fid,
                "name": name,
                "name_kana": "",
                "ward": ward,
                "address": "",
                "lat": "",
                "lng": "",
                "map_url": "",
                "facility_type": "",
                "phone": "",
                "website": "",
                "notes": "",
                "nearest_station": "",
                "station_kana": "",
                "walk_minutes": None,
                "updated": m,
                "totals": {
                    "accept": tot_accept,
                    "wait": tot_wait,
                    "enrolled": tot_enrolled,
                    "capacity_est": cap_est,
                    "wait_per_capacity_est": ratio_opt(tot_wait, cap_est),
                },
                "age_groups": age_groups,
                "ages_0_5": ages_0_5,
            }

            if APPLY_MASTER:
                mm = master.get(fid)
                if mm:
                    injected_cells += apply_master_to_facility(fobj, mm)

            facilities.append(fobj)

        out = {"month": m, "ward": (WARD_FILTER or "横浜市"), "facilities": facilities}
        out_path.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
        print("wrote:", out_path.name, "facilities:", len(facilities), "master_injected_cells:", injected_cells)
        changed_any += 1

    ms = set(existing_months)
    for m in available:
        p = DATA_DIR / f"{m}.json"
        if p.exists() and p.stat().st_size > 200:
            ms.add(m)
    MONTHS_JSON.write_text(json.dumps({"months": sorted(ms)}, ensure_ascii=False, indent=2), encoding="utf-8")
    print("updated months.json:", len(ms), "changed_month_files:", changed_any)


if __name__ == "__main__":
    main()
